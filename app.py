"""
高爾夫球賽管理系統 V1.2

功能說明：
1. 賽事管理
   - 新增/編輯/刪除賽事
   - 賽事日期設定

2. 參賽者管理
   - Excel 檔案匯入參賽者資料
   - 自動產生報名序號（A01、A02...）
   - 支援會員編號、姓名、差點、性別等資料
   - 可手動修改參賽者性別

3. 分組管理
   - 手動拖曳分組
   - 自動分組功能
   - 分組儲存功能
   - 匯出分組表（Excel格式）
   - 女生資料特別標示（粉紅色底色）

4. 報到管理
   - 參賽者報到功能
   - 已報到者不可刪除

版本更新紀錄：
V1.0 - 基礎功能建立
V1.1 - 新增分組匯出功能
V1.2 - 完善分組表格式，加入性別標示
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import os
from datetime import datetime
from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from sqlalchemy import func
from config import config
from extensions import db, init_extensions
import re
import tempfile

# 創建應用程式
app = Flask(__name__)

# 獲取環境配置
env = os.getenv('FLASK_ENV', 'development').strip()
app.config.from_object(config[env])

# 確保實例文件夾存在
if not os.path.exists('instance'):
    os.makedirs('instance')

print(f"數據庫路徑: {app.config['SQLALCHEMY_DATABASE_URI']}")

# 初始化擴展
init_extensions(app)

# 設置 CORS
CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:3000", "https://gold-tawny.vercel.app"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Accept", "Authorization"],
        "supports_credentials": True,
        "max_age": 3600
    }
})

# 設置全局 CORS 響應頭
@app.after_request
def after_request(response):
    origin = request.headers.get('Origin')
    allowed_origins = ["http://localhost:3000", "https://gold-tawny.vercel.app"]
    
    if origin in allowed_origins:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Accept, Authorization'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Max-Age'] = '3600'
    return response

# 處理 OPTIONS 請求
@app.route('/api/v1/tournaments', methods=['OPTIONS'])
def handle_options():
    response = jsonify({'status': 'ok'})
    return response

# 導入模型（在初始化之後）
from models import Tournament, Participant

# 創建所有表
with app.app_context():
    db.create_all()

# 全局函數：解析差點
def parse_handicap(handicap_str):
    try:
        # 如果是數字0，直接返回0.0
        if isinstance(handicap_str, (int, float)) and handicap_str == 0:
            return 0.0
            
        # 如果是空值，返回999.0
        if handicap_str is None or pd.isna(handicap_str) or str(handicap_str).strip() == '' or str(handicap_str).lower() == 'nan':
            return 999.0

        handicap_str = str(handicap_str).strip()
        
        # 如果是純數字，直接轉換
        try:
            return float(handicap_str)
        except ValueError:
            pass
        
        # 嘗試提取括號內的數字
        match = re.search(r'\(([-+]?\d+\.?\d*)\)', handicap_str)
        if match:
            return float(match.group(1))
        
        # 嘗試提取任何數字
        match = re.search(r'([-+]?\d+\.?\d*)', handicap_str)
        if match:
            return float(match.group(1))

        return 999.0
    except Exception as e:
        print(f"解析差點時發生錯誤，輸入值：{handicap_str}，錯誤：{str(e)}")
        return 999.0

# 全局函數：清理文字
def clean_text(text):
    """清理文字，確保是有效的 Unicode 字符"""
    if pd.isna(text):
        return None
    try:
        # 轉換為字符串並移除前後空白
        text = str(text).strip()
        if text == '' or text.lower() == 'nan':
            return None
            
        # 確保文字是有效的 Unicode，並處理任何無效的字符
        text = text.encode('utf-8', errors='ignore').decode('utf-8')
        
        # 移除任何控制字符
        text = ''.join(char for char in text if char.isprintable())
        
        return text
    except Exception as e:
        print(f"清理文字時發生錯誤：{str(e)}")
        return None

# API 路由
@app.route('/api/v1/tournaments', methods=['GET'])
def get_tournaments():
    try:
        print("收到獲取賽事列表請求")
        print(f"請求來源: {request.headers.get('Origin')}")
        print(f"請求方法: {request.method}")
        print(f"請求頭部: {dict(request.headers)}")
        
        tournaments = Tournament.query.all()
        result = []
        for tournament in tournaments:
            result.append({
                'id': tournament.id,
                'name': tournament.name,
                'date': tournament.date.strftime('%Y-%m-%d') if tournament.date else None
            })
        print(f"返回賽事列表: {result}")
        
        response = jsonify(result)
        return response
        
    except Exception as e:
        print(f"獲取賽事列表時發生錯誤: {str(e)}")
        return jsonify({'error': str(e)}), 500

# 建立新賽事
@app.route('/api/v1/tournaments', methods=['POST'])
def create_tournament():
    try:
        print("開始建立新賽事...")
        data = request.json
        print(f"接收到的數據：{data}")
        
        tournament = Tournament(
            name=data['name'],
            date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
            location='',  # 設置為空字符串
            description=''  # 設置為空字符串
        )
        print(f"建立賽事對象：{tournament}")
        
        db.session.add(tournament)
        db.session.commit()
        print(f"賽事保存成功，ID：{tournament.id}")
        
        result = tournament.to_dict()
        print(f"返回結果：{result}")
        return jsonify(result), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"建立賽事時發生錯誤：{str(e)}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# 獲取賽事的參賽者列表
@app.route('/api/v1/tournaments/<int:tournament_id>/participants', methods=['GET'])
def get_tournament_participants(tournament_id):
    try:
        participants = Participant.query.filter_by(tournament_id=tournament_id).order_by(Participant.display_order).all()
        print(f"\n獲取賽事 {tournament_id} 的參賽者列表")
        print(f"總共找到 {len(participants)} 位參賽者")
        
        result = []
        for p in participants:
            participant_dict = p.to_dict()
            print(f"參賽者資料：姓名={p.name}, 預分組編號={p.pre_group_code}")
            result.append(participant_dict)
            
        return jsonify(result)
        
    except Exception as e:
        print(f"獲取參賽者列表時發生錯誤：{str(e)}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# 匯入參賽者
@app.route('/api/v1/tournaments/<int:tournament_id>/participants/import', methods=['POST'])
def import_participants(tournament_id):
    try:
        if 'file' not in request.files:
            return jsonify({'error': '未找到檔案'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未選擇檔案'}), 400
            
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': '請上傳 Excel 檔案 (.xlsx)'}), 400

        # 讀取 Excel 檔案
        df = pd.read_excel(file)
        print(f"Excel 欄位：{df.columns.tolist()}")
        print("Excel 資料預覽：")
        print(df.head())
        
        # 檢查必要欄位
        required_columns = ['姓名', '差點']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'缺少必要欄位：{", ".join(missing_columns)}'}), 400

        # 清除既有的參賽者資料
        Participant.query.filter_by(tournament_id=tournament_id).delete()
        
        # 匯入新的參賽者資料
        for index, row in df.iterrows():
            # 解析差點
            handicap = parse_handicap(row['差點'])
            
            # 解析性別（如果有）
            gender = row.get('性別', '男')
            if pd.isna(gender) or str(gender).strip() == '':
                gender = '男'
            
            # 直接處理預分組編號
            pre_group_code = None
            if '預分組編號' in df.columns:
                raw_value = row['預分組編號']
                print(f"\n第 {index+1} 筆資料的預分組編號原始值：{raw_value}，類型：{type(raw_value)}")
                
                if not pd.isna(raw_value):  # 檢查是否為 NaN
                    if isinstance(raw_value, (int, float)):
                        pre_group_code = str(int(raw_value))  # 轉換數字為字串
                    else:
                        pre_group_code = str(raw_value).strip()  # 其他類型轉換為字串
                    
                    if pre_group_code == '' or pre_group_code.lower() == 'nan':
                        pre_group_code = None
                    
                print(f"處理後的預分組編號：{pre_group_code}")
            
            # 建立參賽者
            participant = Participant(
                tournament_id=tournament_id,
                name=clean_text(str(row['姓名'])),
                gender=gender,
                handicap=handicap,
                member_number=str(row.get('會員編號', '')),
                registration_number=f'A{index+1:02d}',
                pre_group_code=pre_group_code,
                display_order=index
            )
            db.session.add(participant)
        
        db.session.commit()
        return jsonify({'message': '匯入成功'}), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"匯入參賽者時發生錯誤：{str(e)}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# 獲取下一個報名序號
@app.route('/api/v1/tournaments/<int:tournament_id>/next-registration-number', methods=['GET'])
def get_next_registration_number(tournament_id):
    try:
        # 獲取當前賽事的所有參賽者
        participants = Participant.query.filter_by(tournament_id=tournament_id).all()
        
        # 如果沒有參賽者，從 A01 開始
        if not participants:
            return jsonify({'next_number': 'A01'})
            
        # 找出最大的報名序號
        max_number = 0
        for p in participants:
            if p.registration_number and p.registration_number.startswith('A'):
                try:
                    number = int(p.registration_number[1:])
                    max_number = max(max_number, number)
                except ValueError:
                    continue
                    
        # 返回下一個序號
        next_number = f'A{(max_number + 1):02d}'
        return jsonify({'next_number': next_number})
        
    except Exception as e:
        print(f"獲取下一個報名序號時出錯：{str(e)}")
        return jsonify({'error': str(e)}), 500

# 刪除賽事
@app.route('/api/v1/tournaments/<int:tournament_id>', methods=['DELETE'])
def delete_tournament(tournament_id):
    try:
        print(f"開始刪除賽事 ID：{tournament_id}")
        tournament = Tournament.query.get_or_404(tournament_id)
        print(f"找到賽事：{tournament}")
        
        # 先刪除所有相關的參賽者
        print("刪除相關的參賽者")
        Participant.query.filter_by(tournament_id=tournament_id).delete()
        
        # 再刪除賽事本身
        print("刪除賽事本身")
        db.session.delete(tournament)
        
        # 提交事務
        db.session.commit()
        print(f"賽事刪除成功")
        
        return '', 204
        
    except Exception as e:
        db.session.rollback()
        print(f"刪除賽事時發生錯誤：{str(e)}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# 刪除參賽者
@app.route('/api/v1/tournaments/<int:tournament_id>/participants/<int:participant_id>', methods=['DELETE'])
def delete_participant(tournament_id, participant_id):
    try:
        participant = Participant.query.get(participant_id)
        if not participant:
            return jsonify({'error': '找不到指定的參賽者'}), 404
            
        if participant.tournament_id != tournament_id:
            return jsonify({'error': '參賽者不屬於指定的賽事'}), 400
            
        if participant.check_in_status == 'checked_in':
            return jsonify({'error': '已報到的參賽者不能刪除'}), 400
            
        db.session.delete(participant)
        db.session.commit()
        
        return jsonify({'message': '參賽者已成功刪除'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# 更新報到狀態
@app.route('/api/v1/participants/<int:participant_id>/check-in', methods=['PUT'])
def update_check_in_status(participant_id):
    try:
        data = request.json
        check_in_status = data.get('check_in_status')
        check_in_time = data.get('check_in_time')
        
        participant = Participant.query.get(participant_id)
        if not participant:
            return jsonify({'error': '找不到指定的參賽者'}), 404
            
        participant.check_in_status = check_in_status
        if check_in_time:
            participant.check_in_time = datetime.fromisoformat(check_in_time.replace('Z', '+00:00'))
        else:
            participant.check_in_time = None
            
        db.session.commit()
        
        return jsonify({
            'message': '報到狀態更新成功',
            'participant': participant.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"更新報到狀態時發生錯誤：{str(e)}")
        return jsonify({'error': str(e)}), 500

# 自動分組
@app.route('/api/v1/tournaments/<int:tournament_id>/auto-group', methods=['POST'])
def auto_group(tournament_id):
    try:
        # 獲取賽事
        tournament = Tournament.query.get(tournament_id)
        if not tournament:
            return jsonify({'error': '找不到指定的賽事'}), 404

        # 獲取所有參賽者
        participants = Participant.query.filter_by(tournament_id=tournament_id).all()
        if not participants:
            return jsonify({'error': '沒有參賽者可供分組'}), 400

        # 先按預分組編號排序，再按差點排序
        sorted_participants = sorted(participants, 
            key=lambda p: (
                p.pre_group_code if p.pre_group_code else 'Z999',  # 沒有預分組的排最後
                float(p.handicap if p.handicap is not None else 999.0)
            )
        )

        # 計算每組人數（預設 4 人一組）
        group_size = 4
        total_groups = (len(sorted_participants) + group_size - 1) // group_size

        # 進行分組
        for i, participant in enumerate(sorted_participants):
            group_number = (i // group_size) + 1
            participant.group_code = str(group_number)
            participant.display_order = i + 1

        # 儲存變更
        db.session.commit()

        return jsonify({
            'message': '自動分組完成',
            'total_groups': total_groups,
            'total_participants': len(sorted_participants)
        })

    except Exception as e:
        db.session.rollback()
        print('自動分組錯誤:', str(e))
        return jsonify({'error': '自動分組失敗：' + str(e)}), 500

# 儲存分組
@app.route('/api/v1/tournaments/<int:tournament_id>/groups/save', methods=['PUT'])
def save_groups(tournament_id):
    try:
        data = request.json
        groups = data.get('groups', [])
        group_order = data.get('group_order', [])
        
        if not groups:
            return jsonify({'error': '未提供分組資料'}), 400
            
        # 獲取賽事
        tournament = Tournament.query.get(tournament_id)
        if not tournament:
            return jsonify({'error': '找不到指定的賽事'}), 404
            
        # 更新所有參賽者的顯示順序和分組
        display_order = 1
        
        # 按照 group_order 的順序處理各組
        for group_code in group_order:
            group = next((g for g in groups if g['group_code'] == group_code), None)
            if group:
                for participant_id in group['participant_ids']:
                    participant = Participant.query.get(participant_id)
                    if participant and participant.tournament_id == tournament_id:
                        participant.group_code = group_code
                        participant.display_order = display_order
                        display_order += 1
        
        # 處理未分組的參賽者
        unassigned_group = next((g for g in groups if g['group_code'] == '未分組'), None)
        if unassigned_group:
            for participant_id in unassigned_group['participant_ids']:
                participant = Participant.query.get(participant_id)
                if participant and participant.tournament_id == tournament_id:
                    participant.group_code = None
                    participant.display_order = display_order
                    display_order += 1
        
        db.session.commit()
        
        return jsonify({
            'message': '分組儲存成功',
            'total_participants': display_order - 1
        })
        
    except Exception as e:
        db.session.rollback()
        print('儲存分組錯誤:', str(e))
        return jsonify({'error': '儲存分組失敗：' + str(e)}), 500

# 更新分組順序
@app.route('/api/v1/tournaments/<int:tournament_id>/groups/reorder', methods=['PUT'])
def reorder_groups(tournament_id):
    try:
        data = request.get_json()
        group1 = data.get('group1')
        group2 = data.get('group2')

        if not group1 or not group2:
            return jsonify({'error': '缺少組別資訊'}), 400

        # 獲取兩個組別的參賽者
        participants1 = Participant.query.filter_by(
            tournament_id=tournament_id,
            group_code=group1
        ).all()

        participants2 = Participant.query.filter_by(
            tournament_id=tournament_id,
            group_code=group2
        ).all()

        # 交換組別代碼
        for p in participants1:
            p.group_code = group2

        for p in participants2:
            p.group_code = group1

        db.session.commit()

        return jsonify({'message': '組別順序更新成功'})

    except Exception as e:
        db.session.rollback()
        print('更新組別順序錯誤:', str(e))
        return jsonify({'error': '更新組別順序失敗：' + str(e)}), 500

# 更新參賽者組別
@app.route('/api/v1/tournaments/<int:tournament_id>/participants/<int:participant_id>', methods=['PUT'])
def update_participant_group(tournament_id, participant_id):
    try:
        data = request.json
        print(f"接收到的數據：{data}")
        
        participant = Participant.query.get(participant_id)
        if not participant:
            return jsonify({'error': '找不到指定的參賽者'}), 404
            
        if participant.tournament_id != tournament_id:
            return jsonify({'error': '參賽者不屬於指定的賽事'}), 400
            
        target_group = data.get('group_code')
        
        # 更新參賽者組別
        participant.group_code = target_group
        db.session.commit()
        
        print("更新完成")
        return jsonify({
            'message': '更新成功',
            'participant': participant.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print('更新參賽者組別錯誤:', str(e))
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/tournaments/<int:tournament_id>/save_groups', methods=['POST', 'OPTIONS'])
def save_groups_api(tournament_id):
    # 處理 OPTIONS 請求
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
        return response

    try:
        data = request.get_json()
        print(f"接收到的數據: {data}")
        
        if not data or 'groups' not in data:
            return jsonify({'error': '無效的請求資料'}), 400
            
        groups_data = data['groups']
        print(f"接收到的分組數據: {groups_data}")
        
        # 更新所有參賽者的分組
        for group_info in groups_data:
            group_code = group_info['group_code']
            participant_ids = group_info['participant_ids']
            print(f"處理組別 {group_code}, 參賽者: {participant_ids}")
            
            # 更新每個參賽者的分組
            for display_order, participant_id in enumerate(participant_ids, start=1):
                participant = Participant.query.get(participant_id)
                if participant:
                    participant.group_code = group_code
                    participant.display_order = display_order
                    print(f"更新參賽者 {participant_id} 到組別 {group_code}, 順序 {display_order}")
        
        db.session.commit()
        response = jsonify({'message': '分組儲存成功'})
        return response
        
    except Exception as e:
        db.session.rollback()
        print(f"保存分組時發生錯誤：{str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/tournaments/<int:tournament_id>/export_groups', methods=['GET'])
def export_groups(tournament_id):
    try:
        # 獲取賽事信息
        tournament = Tournament.query.get(tournament_id)
        if not tournament:
            return jsonify({'error': '找不到賽事'}), 404

        # 獲取所有參賽者並按分組和顯示順序排序
        participants = Participant.query.filter_by(tournament_id=tournament_id).order_by(
            func.cast(Participant.group_code, db.Integer).asc(),  # 將組別轉換為數字進行排序
            Participant.display_order.asc(),
            Participant.registration_number.asc()
        ).all()

        # 創建一個新的 Excel 工作簿
        wb = openpyxl.Workbook()
        
        # 創建分組名單工作表（放在最前面）
        ws_list = wb.active
        ws_list.title = "分組名單"
        
        # 設置標題
        ws_list.append([f"{tournament.name} 分組名單"])
        ws_list.append(["姓名", "性別", "備註"])
        
        # 設置標題樣式
        title_font = Font(name='微軟正黑體', size=14, bold=True)
        header_font = Font(name='微軟正黑體', size=12, bold=True)
        ws_list['A1'].font = title_font
        ws_list.merge_cells('A1:C1')
        ws_list['A1'].alignment = Alignment(horizontal='center')
        
        for cell in ws_list[2]:
            cell.font = header_font
            
        # 按組別分類參賽者
        current_group = None
        row_idx = 3
        
        for p in participants:
            if p.group_code != current_group:
                current_group = p.group_code
                group_name = f"第 {current_group} 組" if current_group else "未分組"
                ws_list.append([group_name])
                ws_list.merge_cells(f'A{row_idx}:C{row_idx}')
                ws_list[f'A{row_idx}'].font = Font(name='微軟正黑體', size=12, bold=True)
                ws_list[f'A{row_idx}'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                row_idx += 1
            
            # 添加參賽者資料
            gender = "女" if p.gender == "F" else "男"
            ws_list.append([p.name, gender, p.notes or ''])
            
            # 如果是女生，設置粉紅色背景
            if p.gender == "F":
                for cell in ws_list[row_idx]:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            row_idx += 1
        
        # 調整欄寬
        ws_list.column_dimensions['A'].width = 20
        ws_list.column_dimensions['B'].width = 10
        ws_list.column_dimensions['C'].width = 30
        
        # 創建詳細資料工作表
        ws_detail = wb.create_sheet("詳細資料")
        
        # 設置標題
        ws_detail.append([f"{tournament.name} 分組詳細資料"])
        ws_detail.append(["報名序號", "會員編號", "姓名", "差點", "預分組", "分組", "性別", "備註"])
        
        # 設置標題樣式
        ws_detail['A1'].font = title_font
        ws_detail.merge_cells('A1:H1')
        ws_detail['A1'].alignment = Alignment(horizontal='center')
        
        for cell in ws_detail[2]:
            cell.font = header_font
        
        # 添加參賽者資料
        for p in participants:
            gender = "女" if p.gender == "F" else "男"
            ws_detail.append([
                p.registration_number,
                p.member_number,
                p.name,
                p.handicap,
                p.pre_group_code or '',
                p.group_code or '',
                gender,
                p.notes or ''
            ])
            
            # 如果是女生，設置粉紅色背景
            if p.gender == "F":
                row = ws_detail[ws_detail.max_row]
                for cell in row:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # 調整欄寬
        ws_detail.column_dimensions['A'].width = 15
        ws_detail.column_dimensions['B'].width = 15
        ws_detail.column_dimensions['C'].width = 20
        ws_detail.column_dimensions['D'].width = 10
        ws_detail.column_dimensions['E'].width = 10
        ws_detail.column_dimensions['F'].width = 10
        ws_detail.column_dimensions['G'].width = 10
        ws_detail.column_dimensions['H'].width = 30
        
        # 保存到 BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{tournament.name}_分組名單.xlsx'
        )
        
    except Exception as e:
        print(f"匯出分組時發生錯誤：{str(e)}")
        return jsonify({'error': str(e)}), 500

# 匯出分組圖
@app.route('/api/v1/tournaments/<int:tournament_id>/export_groups_diagram', methods=['GET'])
def export_groups_diagram(tournament_id):
    try:
        # 獲取賽事資訊
        tournament = Tournament.query.get_or_404(tournament_id)
        
        # 獲取所有參賽者並按分組和顯示順序排序
        participants = Participant.query.filter_by(tournament_id=tournament_id).order_by(
            func.cast(Participant.group_code, db.Integer).asc(),
            Participant.display_order.asc(),
            Participant.registration_number.asc()
        ).all()

        # 按組別分組
        groups = {}
        for p in participants:
            if p.group_code and p.group_code != '未分組':
                if p.group_code not in groups:
                    groups[p.group_code] = []
                groups[p.group_code].append(p)

        if not groups:
            return jsonify({'error': '沒有已分組的參賽者'}), 400

        # 生成 HTML
        html = '''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>分組圖</title>
            <style>
                body { 
                    font-family: Arial, "Microsoft JhengHei", sans-serif; 
                    padding: 20px;
                    background-color: #f5f5f5;
                }
                .group-container {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 20px;
                    margin-bottom: 20px;
                }
                .group-card {
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    padding: 16px;
                    width: 300px;
                }
                .group-header {
                    margin-bottom: 16px;
                    color: #1976d2;
                    font-size: 1.2em;
                    font-weight: bold;
                }
                .group-code {
                    color: #666;
                    font-size: 0.9em;
                }
                .participant {
                    display: flex;
                    align-items: center;
                    padding: 8px 0;
                    border-bottom: 1px solid #eee;
                }
                .participant:last-child {
                    border-bottom: none;
                }
                .gender-icon {
                    margin: 0 8px;
                    color: #2196f3;
                }
                .gender-icon.female {
                    color: #f06292;
                }
                .handicap {
                    margin-left: auto;
                    color: #666;
                }
                .drag-handle {
                    color: #ccc;
                    margin-right: 8px;
                }
            </style>
        </head>
        <body>
            <div class="group-container">
        '''

        # 添加每個分組的卡片
        for group_code in sorted(groups.keys(), key=lambda x: int(x)):
            group = groups[group_code]
            html += f'''
                <div class="group-card">
                    <div class="group-header">
                        第 {group_code} 組 {len(group)} 人
                        <div class="group-code">預分組: G{int(group_code):02d}</div>
                    </div>
            '''
            
            # 添加組內的參賽者
            for p in group:
                gender_icon = '♀' if p.gender == "F" else '♂'
                gender_class = 'female' if p.gender == "F" else ''
                html += f'''
                    <div class="participant">
                        <span class="drag-handle">≡</span>
                        <span>{p.name}</span>
                        <span class="gender-icon {gender_class}">{gender_icon}</span>
                        <span class="handicap">差點: {p.handicap}</span>
                    </div>
                '''
            
            html += '</div>'

        html += '''
            </div>
        </body>
        </html>
        '''

        # 創建一個臨時文件來保存 HTML
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8') as f:
            f.write(html)
            temp_path = f.name

        response = send_file(
            temp_path,
            mimetype='text/html',
            as_attachment=True,
            download_name=f'{tournament.name}_分組圖.html'
        )

        # 設置 headers 避免快取
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        
        return response

    except Exception as e:
        print(f"匯出分組圖時發生錯誤：{str(e)}")
        return jsonify({'error': str(e)}), 500

# 更新參賽者備註
@app.route('/api/v1/tournaments/<int:tournament_id>/participants/<int:participant_id>/notes', methods=['PUT'])
def update_participant_notes(tournament_id, participant_id):
    try:
        data = request.get_json()
        notes = data.get('notes', '')

        participant = Participant.query.filter_by(
            tournament_id=tournament_id,
            id=participant_id
        ).first_or_404()

        participant.notes = notes
        db.session.commit()

        return jsonify({
            'message': '備註更新成功',
            'participant': participant.to_dict()
        })

    except Exception as e:
        return jsonify({
            'message': f'備註更新失敗: {str(e)}',
            'error': True
        }), 400

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    app.run(host='127.0.0.1', port=port, debug=True)
